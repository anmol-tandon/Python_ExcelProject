"""
Used openpyxl module to read and write excel files
"""
import openpyxl
from openpyxl import Workbook

w_book = openpyxl.load_workbook("RandomData.xlsx")
sheets = w_book.sheetnames
print(sheets)
print(w_book.active.title)

ps_no = input('''Enter ps number from following list
99004319
99004320
99004322
99004324
99004356
99004357
99004358
99004359
99004360
99004361
99004362
99004363
99004364
99004365
99004366
\n ''')

category = input('''Enter for which category of following you want to fetch data
Marks
Hobby
Cities
Languages
Domain \n ''')

if category == 'Marks':
    sh = w_book['Marks']
elif category == 'Hobby':
    sh = w_book['Hobby']
elif category == 'Cities':
    sh = w_book['Cities']
elif category == 'Languages':
    sh = w_book['Languages']
elif category == 'Domain':
    sh = w_book['Domain']
else:
    print("Wrong input")
if ps_no == '99004319':
    BLOCK = 'B2'
elif ps_no == '99004320':
    BLOCK = 'B3'
elif ps_no == '99004322':
    BLOCK = 'B4'
elif ps_no == '99004324':
    BLOCK = 'B5'
elif ps_no == '99004356':
    BLOCK = 'B6'
elif ps_no == '99004357':
    BLOCK = 'B7'
elif ps_no == '99004358':
    BLOCK = 'B8'
elif ps_no == '99004359':
    BLOCK = 'B9'
elif ps_no == '99004360':
    BLOCK = 'B10'
elif ps_no == '99004361':
    BLOCK = 'B11'
elif ps_no == '99004362':
    BLOCK = 'B12'
elif ps_no == '99004363':
    BLOCK = 'B13'
elif ps_no == '99004364':
    BLOCK = 'B14'
elif ps_no == '99004365':
    BLOCK = 'B15'
elif ps_no == '99004366':
    BLOCK = 'B16'
else:
    print("Wrong input")
data = sh[BLOCK].value
# print(data)

w_book2 = Workbook()

w_book2['Sheet'].title = "Final Data"
sheet = w_book2.active
sheet['A1'].value = "Pay Sheet Number"
sheet['A2'].value = ps_no

if category == 'Marks':
    sheet['B1'].value = "Marks"
elif category == 'Hobby':
    sheet['B1'].value = "Hobby"
elif category == 'Cities':
    sheet['B1'].value = "Cities"
elif category == 'Languages':
    sheet['B1'].value = "Languages"
elif category == 'Domain':
    sheet['B1'].value = "Domain"
else:
    print("Wrong input \n")

sheet['B2'].value = data
w_book2.save("Final Output.xlsx")
