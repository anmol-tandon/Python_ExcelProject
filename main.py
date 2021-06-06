import openpyxl

w_book = openpyxl.load_workbook("RandomData.xlsx")
sheets = w_book.sheetnames
print(sheets)
print(w_book.active.title)

ps_no = input('''Enter ps number from following list
99004319
99004320
99004322
99004324
99004356
99004357
99004358
99004359
99004360
99004361
99004362
99004363
99004364
99004365
99004366
\n ''')

category = input('''Enter for which category of following you want to fetch data
Marks
Hobby
Cities
Languages
Domain \n ''')

if category == 'Marks':
    sh = w_book['Marks']
elif category == 'Hobby':
    sh = w_book['Hobby']
elif category == 'Cities':
    sh = w_book['Cities']
elif category == 'Languages':
    sh = w_book['Languages']
elif category == 'Domain':
    sh = w_book['Domain']
else:
    print("Wrong input")
if ps_no == '99004319':
    block = 'B2'
elif ps_no == '99004320':
    block = 'B3'
elif ps_no == '99004322':
    block = 'B4'
elif ps_no == '99004324':
    block = 'B5'
elif ps_no == '99004356':
    block = 'B6'
elif ps_no == '99004357':
    block = 'B7'
elif ps_no == '99004358':
    block = 'B8'
elif ps_no == '99004359':
    block = 'B9'
elif ps_no == '99004360':
    block = 'B10'
elif ps_no == '99004361':
    block = 'B11'
elif ps_no == '99004362':
    block = 'B12'
elif ps_no == '99004363':
    block = 'B13'
elif ps_no == '99004364':
    block = 'B14'
elif ps_no == '99004365':
    block = 'B15'
elif ps_no == '99004366':
    block = 'B16'
else:
    print("Wrong input")
data = sh[block].value
print(data)

from openpyxl import Workbook
w_book2 = Workbook()

w_book2['Sheet'].title = "Final Data"
sheet = w_book2.active
sheet['A1'].value = "Pay Sheet Number"
sheet['A2'].value = ps_no



if category == 'Marks':
    sheet['B1'].value = "Marks"
elif category == 'Hobby':
    sheet['B1'].value = "Hobby"
elif category == 'Cities':
    sheet['B1'].value = "Cities"
elif category == 'Languages':
    sheet['B1'].value = "Languages"
elif category == 'Domain':
    sheet['B1'].value = "Domain"
else:
    print("Wrong input \n")

sheet['B2'].value = data


w_book2.save("Final Output.xlsx")
